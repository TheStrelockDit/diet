from PyQt5.QtWidgets import *
from PyQt5.QtGui import *

from controller import *
import datetime

import pandas as pd
from openpyxl import *
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font


class Abstract(QMainWindow):
    def __init__(self):
        super().__init__()
        self.font = QFont()
        self.font.setFamily('Times New Roman')
        self.font.setPointSize(16)
        self.font.setWeight(2)
        self.font.setBold(True)
        self.font2 = QFont()
        self.font2.setFamily('Times New Roman')
        self.font2.setPointSize(16)
        self.font2.setWeight(1)
        self.font2.setBold(True)
        self.initUI()


class MainMenu(Abstract):
    def initUI(self):
        product = QAction('Продукты', self)
        dishAction = QAction('Блюда', self)
        docAction = QAction('Меню', self)
        menubar = self.menuBar()
        menubar.addAction(product)
        menubar.addAction(dishAction)
        menubar.addAction(docAction)
        menubar.setFont(self.font)
        btn = QPushButton('Требования', self)
        btn.setGeometry(100, 200, 400, 50)
        btn.setFont(self.font)
        labe1 = QLabel('С', self)
        labe1.setGeometry(50, 100, 50, 50)
        labe1.setFont(self.font)
        labe2 = QLabel('До', self)
        labe2.setGeometry(350, 100, 50, 50)
        labe2.setFont(self.font)
        self.date = QDateEdit(self)
        self.date.setGeometry(100, 100, 200, 50)
        self.date.setFont(self.font)
        self.date2 = QDateEdit(self)
        self.date2.setGeometry(400, 100, 200, 50)
        self.date2.setFont(self.font)
        self.date.setDate(datetime.date.today())
        self.date2.setDate(datetime.date.today())
        self.date.hide()
        self.date2.hide()
        labe1.hide()
        labe2.hide()
        btn.hide()
        product.triggered.connect(self.openProductWindow)
        dishAction.triggered.connect(self.openDishWindow)
        docAction.triggered.connect(self.openDocWindow)
        btn.clicked.connect(self.printT)
        self.setGeometry(100, 100, 900, 600)
        self.setWindowTitle('Главное меню')
        self.show()

    def printT(self):
        print(self.date.text())
        self.date.date().setDate(self.date.date().year(), self.date.date().month(), self.date.date().day() + 1)

        print(self.date.text())

    def openProductWindow(self):
        self.p = ProductWin()
        self.destroy()

    def openDishWindow(self):
        self.d = DishWin()
        self.destroy()

    def openDocWindow(self):
        self.d = DocWin()
        self.destroy()


class ProductWin(Abstract):
    def initUI(self):
        mainMenu = QAction('Главное меню', self)
        dishAction = QAction('Блюда', self)
        docAction = QAction('Меню', self)
        self.labelError = QLabel('Ошибка', self)
        self.labelError.setFont(self.font2)
        self.labelError.setStyleSheet("color: rgb(255, 0, 0)")
        self.labelError.setGeometry(200, 50, 300, 30)
        self.labelError.hide()
        menubar = self.menuBar()
        menubar.addAction(mainMenu)
        menubar.addAction(dishAction)
        menubar.addAction(docAction)
        menubar.setFont(self.font)
        label1 = QLabel('Поле ввода', self)
        label1.setFont(self.font)
        label2 = QLabel('Список продуктов', self)
        label2.setFont(self.font)
        self.productList = QComboBox(self)
        self.productList.setFont(self.font)
        btn2 = QPushButton('Удалить', self)
        btn2.setFont(self.font)
        btn3 = QPushButton('Редактировать', self)
        btn3.setFont(self.font)
        self.line = QLineEdit(self)
        self.line.setFont(self.font)
        btn = QPushButton('Добавить', self)
        btn.setFont(self.font)
        label1.setGeometry(50, 100, 200, 50)
        self.line.setGeometry(250, 100, 350, 50)
        btn.setGeometry(650, 100, 200, 50)
        label2.setGeometry(50, 350, 250, 50)
        self.productList.setGeometry(350, 350, 500, 50)
        self.updateList()
        btn2.setGeometry(50, 500, 350, 50)
        btn3.setGeometry(500, 500, 350, 50)
        self.productList.setEditable(True)

        mainMenu.triggered.connect(self.openMainWindow)
        dishAction.triggered.connect(self.openDishWindow)
        docAction.triggered.connect(self.openDocWindow)

        btn.clicked.connect(self.addNewProduct)
        btn2.clicked.connect(self.deleteNewProduct)
        btn3.clicked.connect(self.redactNewProduct)

        self.setGeometry(100, 100, 900, 600)
        self.setWindowTitle('Продукты')
        self.show()

    def addNewProduct(self):
        self.labelError.hide()
        newProduct = self.line.text().lower()
        if newProduct != '' and newProduct != ' ':
            p = Product(newProduct)
            error = p.addNewProduct()
            if error == True:
                self.labelError.setText('Такой продукт уже есть')
                self.labelError.show()
            p.f.saveFile()
            del p, newProduct, error
            self.line.clear()
            self.updateList()

    def redactNewProduct(self):
        self.labelError.hide()
        product = self.productList.currentText()
        self.line.setText(product)
        self.deleteNewProduct()
        del product

    def deleteNewProduct(self):
        self.labelError.hide()
        product = self.productList.currentText()
        p = Product(product)
        p.delNewProduct()
        p.f.saveFile()
        del p, product
        self.updateList()

    def updateList(self):

        f = File()
        self.productList.clear()
        self.productList.addItem('')
        f.list['Продукты'].sort()
        self.productList.addItems(f.list['Продукты'])
        del f

    def openMainWindow(self):
        self.m = MainMenu()
        self.destroy()

    def openDishWindow(self):
        self.d = DishWin()
        self.destroy()

    def openDocWindow(self):
        self.d = DocWin()
        self.destroy()


class DishWin(Abstract):
    def initUI(self):
        self.labelError = QLabel('Ошибка', self)
        self.labelError.setFont(self.font2)
        self.labelError.setStyleSheet("color: rgb(255, 0, 0)")
        self.labelError.setGeometry(200, 50, 500, 30)
        self.labelError.hide()
        product = QAction('Продукты', self)
        mainMenu = QAction('Главное меню', self)
        docAction = QAction('Меню', self)
        label1 = QLabel('Название блюда', self)
        label1.setFont(self.font)
        label2 = QLabel('Список блюд', self)
        label2.setFont(self.font)
        label3 = QLabel('Продукты', self)
        label3.setFont(self.font)
        label4 = QLabel('Список продуктов блюда', self)
        label4.setFont(self.font)
        label5 = QLabel('Вес', self)
        label5.setFont(self.font)
        label6 = QLabel('Ид', self)
        label6.setFont(self.font)
        btn1 = QPushButton('Удалить', self)
        btn1.setFont(self.font)
        btn2 = QPushButton('Создать', self)
        btn2.setFont(self.font)
        btn3 = QPushButton('Добавить', self)
        btn3.setFont(self.font)
        btn4 = QPushButton('Удалить', self)
        btn4.setFont(self.font)
        self.line1 = QLineEdit(self)
        self.line1.setFont(self.font)
        self.line2 = QLineEdit(self)
        self.line2.setFont(self.font)
        self.line3 = QLineEdit(self)
        self.line3.setFont(self.font)
        self.combo1 = QComboBox(self)
        self.combo1.setFont(self.font)
        self.combo2 = QComboBox(self)
        self.combo2.setFont(self.font)
        self.combo1.setEditable(True)
        self.combo2.setEditable(True)
        self.listW = QListWidget(self)
        self.listW.setFont(self.font)
        menubar = self.menuBar()
        menubar.addAction(product)
        menubar.addAction(mainMenu)
        menubar.addAction(docAction)
        menubar.setFont(self.font)
        self.updateDishes()
        self.updateList()

        product.triggered.connect(self.openProductWindow)
        mainMenu.triggered.connect(self.openMainWindow)
        docAction.triggered.connect(self.openDocWindow)
        btn1.clicked.connect(self.removeProductFromDish)
        btn2.clicked.connect(self.createNewDish)
        btn3.clicked.connect(self.addProductToDish)
        btn4.clicked.connect(self.deleteDish)
        self.combo1.currentTextChanged.connect(self.updateListW)

        self.setGeometry(100, 100, 900, 650)
        label1.setGeometry(200, 100-50, 300, 50)
        label2.setGeometry(50, 300-100, 170, 50)
        label3.setGeometry(650, 400-130, 200, 50)
        label4.setGeometry(50, 400-130, 350, 50)
        label5.setGeometry(570, 530-130, 150, 50)
        label6.setGeometry(50, 100-50, 50, 50)
        btn1.setGeometry(50, 700-150, 500, 50)
        btn2.setGeometry(200, 200-80, 500, 50)
        btn3.setGeometry(570, 600-130, 300, 50)
        btn4.setGeometry(700, 300-100, 120, 50)
        self.line1.setGeometry(100, 100-50, 80, 50)
        self.line2.setGeometry(420, 100-50, 400, 50)
        self.line3.setGeometry(650, 530-130, 220, 50)
        self.combo1.setGeometry(250, 300-100, 400, 50)
        self.combo2.setGeometry(570, 450-130, 300, 50)
        self.listW.setGeometry(50, 470-150, 500, 210)

        self.setWindowTitle('Блюда')
        self.show()

    def updateDishes(self):
        self.labelError.hide()
        self.combo1.clear()
        self.combo1.addItem('')
        f = File()
        for id in f.list['Блюда']:
            text = str(id) + '-' + f.list['Блюда'][id]['Название']
            self.combo1.addItem(text)
        del f

    def updateList(self):

        f = File()
        self.combo2.clear()
        self.combo2.addItem('')
        f.list['Продукты'].sort()
        self.combo2.addItems(f.list['Продукты'])
        del f

    def deleteDish(self):
        try:
            id = self.combo1.currentText().split('-')
            name = id[1]
            id = str(id[0])
            d = Dish(name, id)
            d.delNewDish()
            d.f.saveFile()
            self.updateDishes()
            del id, name, d
        except:
            self.labelError.setText('Выберете блюдо, если оно есть!')
            self.labelError.show()

    def updateListW(self):
        try:
            self.labelError.hide()
            self.listW.clear()
            dish = self.combo1.currentText().split('-')
            name = dish[1]
            id = dish[0]

            d = Dish(name, id)
            for i in d.f.list['Блюда'][id]['Продукты']:
                try:
                    weight = d.f.list['Блюда'][id]['Продукты'][i]
                    text = i + '-' + weight
                    self.listW.addItem(text)

                except:
                    pass
            del dish, name, id, d
        except:
            pass

    def addProductToDish(self):

        dish = self.combo1.currentText().split('-')
        name = dish[1]
        id = dish[0]

        product = self.combo2.currentText()
        f = File()
        if product in f.list['Продукты']:

            weight = self.line3.text()
            if weight != '' and ' ':
                d = Dish(name, id)
                d.addProductToDish(product, weight)
                d.f.saveFile()
                self.updateListW()
                del d
            else:
                self.labelError.setText('Введите вес продукта!')
                self.labelError.show()

            self.line3.clear()
            del dish, name, id, product, weight
            self.updateList()
        else:
            self.line3.clear()
            self.combo2.setEditText('')

    def createNewDish(self):
        id = self.line1.text().lower()
        name = self.line2.text().lower()
        if id != '' and id != ' ' and name != '' and name != ' ':
            self.line1.clear()
            self.line2.clear()
            d = Dish(name, id)
            check = d.addNewDish()

            d.f.saveFile()
            self.updateDishes()
            if check is True:
                self.labelError.setText('Блюдо с таким идом уже есть!')
                self.labelError.show()
            del d, id, name, check
        else:
            self.labelError.show()
            self.labelError.setText('Введите ид и название блюда!')
            self.line1.clear()
            self.line2.clear()

    def removeProductFromDish(self):
        try:
            text = self.listW.currentItem().text().split('-')
            text = text[0]
            dish = self.combo1.currentText().split('-')
            id = dish[0]
            f = File()
            del f.list['Блюда'][id]['Продукты'][text]
            f.saveFile()
            self.updateListW()
            del text, dish, id, f
        except:
            self.labelError.setText('Выберете продукт для удаления!')
            self.labelError.show()

    def openProductWindow(self):
        self.p = ProductWin()
        self.destroy()

    def openMainWindow(self):
        self.m = MainMenu()
        self.destroy()

    def openDocWindow(self):
        self.d = DocWin()
        self.destroy()


class DocWin(Abstract):
    def initUI(self):
        product = QAction('Продукты', self)
        dish = QAction('Блюда', self)
        mainMenu = QAction('Главное меню', self)
        self.date = QDateEdit(self)
        self.date.setGeometry(50, 100, 200, 50)
        self.date.setFont(self.font)
        self.date2 = QDateEdit(self)
        self.date2.setGeometry(600, 550, 250, 50)
        self.date2.setFont(self.font)

        labe0 = QLabel('Календарь', self)
        labe0.setFont(self.font)
        labe0.setGeometry(50, 50, 150, 50)
        labe1 = QLabel('Список меню', self)
        labe1.setFont(self.font)
        labe1.setGeometry(50, 200, 200, 50)
        self.listW = QListWidget(self)
        self.listW.setFont(self.font)
        self.listW.setGeometry(50, 250, 500, 370)
        labe2 = QLabel('Название меню', self)
        labe2.setFont(self.font)
        labe2.setGeometry(400, 50, 200, 50)
        self.line1 = QLineEdit(self)
        self.line1.setFont(self.font)
        self.line1.setGeometry(400, 100, 450, 50)
        self.line2 = QLineEdit(self)
        self.line2.setFont(self.font)
        self.line2.setGeometry(50, 650, 500, 50)
        self.line2.hide()
        btn1 = QPushButton('Добавить', self)
        btn1.setFont(self.font)
        btn1.setGeometry(400, 170, 450, 50)
        btn2 = QPushButton('Открыть', self)
        btn2.setFont(self.font)
        btn2.setGeometry(600, 250, 250, 50)
        btn3 = QPushButton('Скопировать', self)
        btn3.setFont(self.font)
        btn3.setGeometry(600, 350, 250, 50)
        btn4 = QPushButton('Удалить', self)
        btn4.setFont(self.font)
        btn4.setGeometry(600, 450, 250, 50)
        self.btn5 = QPushButton('Выбрать', self)
        self.btn5.setGeometry(600, 650, 250, 50)
        self.btn5.setFont(self.font)
        self.date2.hide()
        self.btn5.hide()
        self.date.setDate(datetime.date.today())
        self.date2.setDate(datetime.date.today())
        menubar = self.menuBar()
        menubar.addAction(product)
        menubar.addAction(dish)
        menubar.addAction(mainMenu)
        menubar.setFont(self.font)
        btn2.clicked.connect(self.openDocWinTwo)
        product.triggered.connect(self.openProductWindow)
        dish.triggered.connect(self.openDishWindow)
        mainMenu.triggered.connect(self.openMainWindow)
        btn1.clicked.connect(self.createNewMenu)
        btn4.clicked.connect(self.deleteMenu)
        btn3.clicked.connect(self.showBut)
        self.btn5.clicked.connect(self.copyOut)
        self.date.dateChanged.connect(self.showListW)
        self.setGeometry(100, 100, 900, 750)
        self.setWindowTitle('Документ')
        self.showListW()
        self.show()

    def deleteMenu(self):
        date = self.date.text().lower()
        name = self.listW.currentItem().text()
        m = Menu(name, date)
        m.delNewMenu()
        m.f.saveFile()
        del date, name, m
        self.showListW()

    def showBut(self):
        self.btn5.show()
        self.date2.show()
        self.line2.show()
        self.line2.setText(self.listW.currentItem().text())

    def copyOut(self):
        date = self.date.text().lower()
        name = self.listW.currentItem().text()
        newName = self.line2.text()

        newDate = self.date2.text().lower()
        m = Menu(name, date)
        m.copyMenu(newDate, newName)
        m.f.saveFile()
        del date, name, newName, m, newDate
        self.showListW()
        self.btn5.hide()
        self.date2.hide()
        self.line2.hide()

    def createNewMenu(self):
        date = self.date.text().lower()
        name = self.line1.text().lower()
        m = Menu(name, date)
        m.addNewMenu()
        m.f.saveFile()
        del date, name, m
        self.line1.clear()
        self.showListW()

    def showListW(self):
        self.listW.clear()
        date = self.date.text().lower()
        name = self.line1.text().lower()
        m = Menu(date, name)
        if date in m.f.list['Меню']:
            self.listW.addItems(m.f.list['Меню'][date])
        del date, name, m

    def openDocWinTwo(self):
        date = self.date.text().lower()
        name = self.listW.currentItem().text().lower()
        self.d = DocWinTwo(date, name)
        self.destroy()

    def openProductWindow(self):
        self.p = ProductWin()
        self.destroy()

    def openDishWindow(self):
        self.d = DishWin()
        self.destroy()

    def openMainWindow(self):
        self.m = MainMenu()
        self.destroy()


class DocWinTwo(Abstract):
    def __init__(self, date, name):
        self.date = date
        self.name = name
        super().__init__()

    def initUI(self):
        product = QAction('Продукты', self)
        dish = QAction('Блюда', self)
        mainMenu = QAction('Меню', self)
        self.cb = QComboBox(self)
        self.cb.setFont(self.font)
        self.cb.setGeometry(50, 80, 250, 50)
        self.cb.addItem('Завтрак')
        self.cb.addItem('Второй завтрак')
        self.cb.addItem('Обед')
        self.cb.addItem('Полдник')
        self.cb.addItem('Ужин')
        self.listW = QListWidget(self)
        self.listW.setFont(self.font)
        self.listW.setGeometry(50, 150, 700, 450)
        labe1 = QLabel('Список блюд', self)
        labe1.setFont(self.font)
        labe1.setGeometry(600, 30, 200, 50)
        self.cb2 = QComboBox(self)
        self.cb2.setFont(self.font)
        self.cb2.setGeometry(600, 80, 400, 50)
        self.cb2.setEditable(True)

        self.updateDishes()
        labe2 = QLabel('Порций', self)
        labe2.setFont(self.font)
        labe2.setGeometry(1050, 30, 100, 50)
        self.line1 = QLineEdit(self)
        self.line1.setFont(self.font)
        self.line1.setGeometry(1050, 80, 70, 50)
        btn1 = QPushButton('Сохранить', self)
        btn1.setFont(self.font)
        btn1.setGeometry(800, 250, 320, 50)
        btn2 = QPushButton('Удалить', self)
        btn2.setFont(self.font)
        btn2.setGeometry(800, 350, 320, 50)
        btn3 = QPushButton('Печать', self)
        btn3.setFont(self.font)
        btn3.setGeometry(800, 450, 320, 50)
        menubar = self.menuBar()
        menubar.addAction(product)
        menubar.addAction(dish)
        menubar.addAction(mainMenu)
        menubar.setFont(self.font)
        self.updateDishesType()
        product.triggered.connect(self.openProductWindow)
        dish.triggered.connect(self.openDishWindow)
        mainMenu.triggered.connect(self.openDocWindow)
        btn3.clicked.connect(self.print3)
        self.cb.currentTextChanged.connect(self.updateDishesType)
        btn2.clicked.connect(self.delDish)
        btn1.clicked.connect(self.addDish)
        self.setGeometry(50, 50, 1200, 650)
        self.setWindowTitle('Создание меню')
        self.show()

    def addDish(self):
        type = self.cb.currentText()
        dish = self.cb2.currentText()
        num = self.line1.text()
        dish = dish + '-' + '.......' + '-' + num
        f = File()
        check = True
        if dish != '':
            try:
                p = dish.split('-')
                id = p[0]
                for i in f.list['Меню'][self.date][self.name][type]:
                    pi = i.split('-')
                    idi = pi[0]
                    if id == idi:
                        f.list['Меню'][self.date][self.name][type].remove(i)
                        f.list['Меню'][self.date][self.name][type].append(dish)
                        check = False

                if check is True:
                    f.list['Меню'][self.date][self.name][type].append(dish)
            except:
                f.list['Меню'][self.date][self.name][type] = []
                f.list['Меню'][self.date][self.name][type].append(dish)
            f.saveFile()
            self.updateDishesType()
            self.updateDishes()

    def delDish(self):
        type = self.cb.currentText()
        text = self.listW.currentItem().text()
        f = File()
        f.list['Меню'][self.date][self.name][type].remove(text)
        f.saveFile()
        self.updateDishesType()

    def print3(self):
        wb = Workbook()
        wb.save('test.xlsx')

        wb = load_workbook('test.xlsx')

        # Размеры таблиц
        ws = wb.active
        ws.column_dimensions['A'].width = 50
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        product_list = []
        weight_list = []
        param = 0
        param += 2
        # Заполнение строк
        ws['A' + str(param)].value = self.name
        ws['A' + str(param)].font = Font(italic=True)
        ws['C' + str(param)].value = self.date
        ws['C' + str(param)].font = Font(italic=True)

        df_list = []
        f = File()
        for i in f.list['Меню'][self.date][self.name]:
            param += 2
            ws['A' + str(param)].value = i
            ws['A' + str(param)].font = Font(italic=True)
            for j in f.list['Меню'][self.date][self.name][i]:
                product = []
                kg = []
                exit = []
                text = j.split('-')
                por = float(text[3])
                param += 1
                ws['A' + str(param)].value = text[1]
                ws['A' + str(param)].font = Font(italic=True)
                ws['B' + str(param)].value = 'Порций всего'
                ws['B' + str(param)].font = Font(italic=True)
                ws['C' + str(param)].value = str(int(por))
                ws['C' + str(param)].font = Font(italic=True)
                id = text[0]
                for k in f.list['Блюда'][id]['Продукты']:

                    product.append(k)
                    x = por * float(f.list['Блюда'][id]['Продукты'][k])
                    x = round(x, 5)
                    x = str(x)
                    if k in product_list:
                        ind = product_list.index(k)

                        weight_list[ind] = str(round(float(weight_list[ind]) + float(x), 5))
                    else:
                        product_list.append(k)
                        weight_list.append(x)
                    exit.append(x)
                    kg.append(f.list['Блюда'][id]['Продукты'][k])
                data = {'Продукт': pd.Series(product),
                        'Порция на 1': pd.Series(kg),
                        'Всего': pd.Series(exit)}
                df = pd.DataFrame(data)
                print(data)

                param += df.shape[0] + 1
                for r in dataframe_to_rows(df, index=False):
                    ws.append(r)
                param += 1

        param += 10

        ws['A' + str(param)].value = 'Требования'
        ws['A' + str(param)].font = Font(italic=True)
        ws['B' + str(param)].value = self.name
        ws['B' + str(param)].font = Font(italic=True)
        ws['C' + str(param)].value = self.date
        ws['C' + str(param)].font = Font(italic=True)
        param += 1
        ws['A' + str(param)].value = ''

        data = {'Продукт': pd.Series(product_list),
                'Всего': pd.Series(weight_list)}
        df = pd.DataFrame(data)
        df.sort_values(['Продукт', 'Всего'], ascending=True, inplace=True)
        param += df.shape[0] + 1

        for r in dataframe_to_rows(df, index=False):

            ws.append(r)

        param += 1

        wb.save('test.xlsx')

    def updateDishesType(self):
        type = self.cb.currentText()
        f = File()
        self.listW.clear()
        self.line1.clear()
        try:

            self.listW.addItems(f.list['Меню'][self.date][self.name][type])
        except:
            pass

    def updateDishes(self):
        self.cb2.clear()
        self.cb2.addItem('')
        f = File()
        for id in f.list['Блюда']:
            text = str(id) + '-' + f.list['Блюда'][id]['Название']
            self.cb2.addItem(text)
        del f

    def openProductWindow(self):
        self.p = ProductWin()
        self.destroy()

    def openDishWindow(self):
        self.d = DishWin()
        self.destroy()

    def openDocWindow(self):
        self.d = DocWin()
        self.destroy()
